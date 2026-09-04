#!/usr/bin/env python3
"""query_arbitrators.py — Tier 1 arbitrator analytics over the UNCTAD ISDS
Navigator full-data Excel (31/12/2023 snapshot).

Part of the isds-arbitrators skill (sibling to isds-research).
Data basis: treaty-based ISDS only; commercial-arbitration appointments and
confidential cases are invisible. Every output carries a DATA FRESHNESS footer.
The Excel itself must be the user's own download and never ships.

Usage examples:
  python3 query_arbitrators.py --arbitrator "Stern"
  python3 query_arbitrators.py --arbitrator "Stern" --json
  python3 query_arbitrators.py --list-matches "van Houtte"
  python3 query_arbitrators.py --shortlist --issue "Fair and equitable" \
      --rules ICSID --role President --top 10
  python3 query_arbitrators.py --shortlist --issue expropriation --found \
      --weights my-weights.json
"""
import argparse, json, os, re, sys, unicodedata
from collections import Counter, defaultdict

HEADER_ROW = 12  # headers at row 12, data from 13 (1-indexed)
SNAPSHOT = "31/12/2023"
SNAPSHOT_CASES = 1332
LAST_KNOWN_NAVIGATOR = {"as_of": "31/12/2025", "total": 1463, "observed": "2026-07-02"}

COL = dict(no=0, year=1, short=2, full=3, iia=4, rules=5, inst=6, status=7,
           respondent=8, home=9, sector=10, subsector=11, summary=12,
           investment=13, arbitrators=14, decisions=15, opinions=16,
           claimed=17, awarded=18, alleged=19, found=20, fo_type=21,
           fo_status=22, fo_decisions=23, fo_opinions=24, annulment=25,
           italaw=26, background=27)

PLACEHOLDER_RE = re.compile(r"not available|not constituted|^none$|^n/?a$", re.I)

# R1 (Green, 2026-08-02): the opinions denominator counts only cases with at
# least one SUBSTANTIVE decision — an award, or a decision on jurisdiction,
# the merits/liability, or damages/quantum. Procedural orders, bifurcation,
# rectification, revision, discontinuance orders etc. do NOT count.
SUBSTANTIVE_INC = re.compile(
    r"\baward\b|decision (?:of the tribunal )?on (?:the )?(?:objections to )?"
    r"(?:jurisdiction|admissibilit|liabilit|merits?|quantum|damages?|"
    r"responsibilit|preliminary objections?)", re.I)
# Consent awards excluded per Green 2026-08-02: the arbitrators have little
# say in a consent award, so it is not a dissentable decision.
SUBSTANTIVE_EXC = re.compile(
    r"^order\b|discontinuan|bifurcation|rectification|revision|"
    r"provisional measures|stay of|settlement agreement$|consent award|"
    r"award embodying the parties.{0,3} settlement", re.I)


def classify_opinion_kind(kind):
    k = kind.lower()
    if "concurring" in k and "dissent" in k:
        return "concurring-and-dissenting"
    if "partial" in k and "dissent" in k:
        return "partial dissent"
    if "dissent" in k:
        return "dissent"
    if "separate" in k:
        return "separate"
    if "concurring" in k:
        return "concurring"
    if "declaration" in k or "statement" in k:
        return "declaration"
    return "other"


def substantive_decisions(decisions_field):
    """Return the list of substantive decision titles in a DECISIONS cell."""
    out = []
    v = str(decisions_field or "").strip()
    if not v or v in ("None", "Data not available", "No decisions recorded"):
        return out
    for entry in re.split(r";\s*\n|;(?=[A-Z])", v):
        e = entry.strip()
        if not e:
            continue
        title = re.split(r"\s+dated\s+", e)[0].strip()
        if SUBSTANTIVE_EXC.search(title):
            continue
        if SUBSTANTIVE_INC.search(title):
            out.append(title)
    return out
OPINION_KIND_RE = re.compile(
    r"^(?P<kind>.*?(?:Opinion|Declaration|Statement|Dissent))s?\b(?P<mid>[^;]*?)"
    r"(?:\s+(?:by|of)\s+(?:the\s+arbitrator\s+)?(?P<author>[A-ZÀ-Þ].*?))?\s*"
    r"(?:\((?P<doc>[^)]*)\))?\s*$")
TRAILING_TOPIC_RE = re.compile(r"\s+on\s+(?:one\s+aspect.*|[A-Z][\w &/-]*)$")


def norm(s):
    s = unicodedata.normalize("NFD", s or "")
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).strip().lower()


def token_parts(token):
    """'van den Berg, A. J.' -> (surname_norm, initials) e.g. ('van den berg','AJ')."""
    if "," in token:
        surname, rest = token.split(",", 1)
        initials = "".join(re.findall(r"[A-Z]", rest))
    else:
        surname, initials = token, ""
    return norm(surname), initials


class Dataset:
    def __init__(self, excel_path, overlay):
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        self.rows = []
        for r in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
            if r[COL["short"]]:
                self.rows.append(r)
        self.overlay = overlay
        self.merge = overlay.get("merge", {})
        self.author_overrides = {norm(k): v for k, v in overlay.get("author_overrides", {}).items()}
        self._build()

    def canon(self, token):
        token = re.sub(r"\s+", " ", token.strip())
        return self.merge.get(token, token)

    def _build(self):
        self.seats = []            # dicts: case_idx, token, role
        self.tribunal_by_case = defaultdict(list)
        for i, r in enumerate(self.rows):
            raw = str(r[COL["arbitrators"]] or "").strip()
            if not raw:
                continue
            for part in raw.split(";"):
                part = part.strip()
                if not part:
                    continue
                m = re.match(r"^(?P<name>.+?)\s+-\s+(?P<role>[A-Za-z .()/-]+)$", part)
                if not m:
                    continue
                tok = self.canon(m.group("name"))
                if PLACEHOLDER_RE.search(tok):
                    tok = "Name not available"
                role = m.group("role").strip()
                self.seats.append(dict(case=i, token=tok, role=role))
                self.tribunal_by_case[i].append((tok, role))
        # opinions
        self.opinions = defaultdict(list)   # case_idx -> [(kind, author_raw, token_or_None)]
        self.unattributed = []
        for i, r in enumerate(self.rows):
            raw = str(r[COL["opinions"]] or "").strip()
            if not raw or raw in ("None", "Data not available", "No decisions recorded"):
                continue
            for entry in raw.split(";"):
                entry = entry.strip()
                if not entry:
                    continue
                m = OPINION_KIND_RE.match(entry)
                if not m:
                    continue
                kind = m.group("kind").strip()
                author = (m.group("author") or "").strip()
                # author may be missing from group when phrased 'Opinion on X of NAME'
                if not author and m.group("mid"):
                    m2 = re.search(r"(?:by|of)\s+(?:the\s+arbitrator\s+)?([A-ZÀ-Þ][^(;]+)$",
                                   m.group("mid").strip())
                    if m2:
                        author = m2.group(1).strip()
                def clean(a):
                    a = a.split("(")[0]
                    a = TRAILING_TOPIC_RE.sub("", a).strip()
                    a = re.sub(r"\b(Professor|Prof\.?|Dr\.?|Judge|Sir|Dame|Q\.?C\.?|K\.?C\.?)\b\.?",
                               "", a).strip(" ,")
                    return a
                candidates_a = []
                if author:
                    candidates_a.append(clean(author))
                mid = (m.group("mid") or "").strip()
                if mid.startswith(","):
                    candidates_a.append(clean(mid.lstrip(",").split("(")[0]))
                author, tok = "", None
                for a in candidates_a:
                    if not a:
                        continue
                    t = self._match_author(i, a)
                    if t is not None:
                        author, tok = a, t
                        break
                if not author and candidates_a:
                    author = next((a for a in candidates_a if a), "")
                # R2: recover the decision-document and language attributes from
                # the parenthetical chunks (handles one level of nesting)
                chunks = re.findall(r"\(([^()]*(?:\([^()]*\)[^()]*)*)\)", entry)
                LANGS = ("English", "Spanish", "French", "Russian", "Portuguese",
                         "German", "Italian", "Arabic", "Turkish")
                language = next((c.strip() for c in chunks if c.strip() in LANGS), "")
                decision = next(
                    (re.sub(r"\s*\((?:%s)\)\s*" % "|".join(LANGS), " ", c).strip()
                     for c in chunks
                     if re.search(r"\bdated\b|\baward\b|\bdecision\b", c, re.I)),
                    "")
                self.opinions[i].append(dict(
                    kind=kind, kind_class=classify_opinion_kind(kind),
                    author=author or "(unnamed)", token=tok,
                    decision=decision or "(decision not specified)",
                    language=language))
                if author and tok is None:
                    self.unattributed.append((self.rows[i][COL["short"]], author))
        # annulment committee membership
        self.annulment_by_token = Counter()
        for i, r in enumerate(self.rows):
            raw = str(r[COL["annulment"]] or "").strip()
            if not raw or PLACEHOLDER_RE.search(raw):
                continue
            for part in raw.split(";"):
                part = part.split(" - ")[0].strip()
                if part and not PLACEHOLDER_RE.search(part):
                    self.annulment_by_token[self.canon(part)] += 1

    def _match_author(self, case_idx, author):
        a = norm(author)
        if a in self.author_overrides:
            return self.author_overrides[a]
        candidates = self.tribunal_by_case.get(case_idx, [])
        for tok, _ in candidates:
            sn, _ini = token_parts(tok)
            if a.endswith(sn) or sn in a:
                return tok
        # fall back: search all tokens (opinion rows with empty tribunal field)
        hits = {s["token"] for s in self.seats
                if token_parts(s["token"])[0] and
                (a.endswith(token_parts(s["token"])[0]) or token_parts(s["token"])[0] in a)}
        return hits.pop() if len(hits) == 1 else None

    # ---------- resolution ----------
    def resolve(self, query):
        q = norm(query)
        toks = sorted({s["token"] for s in self.seats})
        exact = [t for t in toks if norm(t) == q]
        if exact:
            return exact
        return [t for t in toks if q in norm(t)]

    # ---------- profile ----------
    def profile(self, token):
        seats = [s for s in self.seats if s["token"] == token]
        cases = sorted({s["case"] for s in seats})
        roles = Counter(s["role"] for s in seats)
        appointing = Counter()
        for s in seats:
            base = s["role"].replace(" (replaced)", "")
            appointing[base] += 1
        # R2/F3: dedup on (case, decision, kind_class); language is an attribute
        opinions_by_me = []
        for c in cases:
            groups = {}
            for o in self.opinions.get(c, []):
                if o["token"] != token:
                    continue
                key = (o["kind_class"], o["decision"])
                g = groups.setdefault(key, dict(
                    case=self.rows[c][COL["short"]], kind=o["kind"],
                    kind_class=o["kind_class"], decision=o["decision"],
                    author=o["author"], languages=[]))
                if o["language"] and o["language"] not in g["languages"]:
                    g["languages"].append(o["language"])
            opinions_by_me.extend(groups.values())
        # R1 denominator: cases with >=1 substantive decision AND opinions data known
        opinions_known = sum(
            1 for c in cases
            if str(self.rows[c][COL["opinions"]] or "").strip() != "Data not available"
            and substantive_decisions(self.rows[c][COL["decisions"]]))
        def col_counter(col):
            ctr = Counter()
            for c in cases:
                v = str(self.rows[c][col] or "").strip()
                if v and not PLACEHOLDER_RE.search(v):
                    ctr[v] += 1
            return ctr
        issue_alleged, issue_found = Counter(), Counter()
        # v0.9.3 (A4/F12): renamed from dissent_in_found — it counts
        # CO-OCCURRENCE (an individual opinion of any kind in a case where that
        # breach was found), never a dissent on that issue. Counting widened
        # from dissent-class to any published individual opinion to match the
        # new name; the rendered table prints the disclaimer with the number.
        opinion_in_found_cases = Counter()
        for c in cases:
            found_items = [p.strip() for p in str(self.rows[c][COL["found"]] or "").split(";") if p.strip()]
            alleged_items = [p.strip() for p in str(self.rows[c][COL["alleged"]] or "").split(";") if p.strip()]
            me_opinion = any(o["token"] == token for o in self.opinions.get(c, []))
            for it in alleged_items:
                if not re.match(r"Pending|Not applicable|None|Data not", it):
                    issue_alleged[it] += 1
            for it in found_items:
                if not re.match(r"Pending|Not applicable|None|Data not", it):
                    issue_found[it] += 1
                    if me_opinion:
                        opinion_in_found_cases[it] += 1
        co = Counter()
        presidents = Counter()
        # v0.9.3 (A6/TC2/TC3): "Presidents sat under" counts only cases where
        # the subject held a WING seat (Respondent/Claimant, incl. replaced).
        # The old all-cases loop listed the presidents who SUCCEEDED a replaced
        # president, and listed peers for subjects who never held a wing seat.
        wing_cases = {s["case"] for s in seats
                      if s["role"].replace(" (replaced)", "").strip()
                      in ("Respondent", "Claimant")}
        for c in cases:
            for tok, role in self.tribunal_by_case[c]:
                if tok != token and tok != "Name not available":
                    co[tok] += 1
                    if role.startswith("President") and c in wing_cases:
                        presidents[tok] += 1
        # v0.9.3 (A10/G4): opinions recorded in the dataset that name (or
        # fallback-match) this arbitrator in cases NOT attributed to them in the
        # ARBITRATORS column (e.g. "Data not available" rows) are invisible to
        # the counts above. Surface them as a verify-first footnote.
        sn, _ini = token_parts(token)
        my_cases = set(cases)
        stray_opinions = []
        for ci, ops in self.opinions.items():
            if ci in my_cases:
                continue
            for o in ops:
                if o["token"] == token or (sn and sn in norm(o["author"])):
                    stray_opinions.append(dict(
                        case=self.rows[ci][COL["short"]], kind=o["kind"],
                        author=o["author"], decision=o["decision"]))
        pending = sum(1 for c in cases if str(self.rows[c][COL["status"]] or "").strip() == "Pending")
        years = [int(self.rows[c][COL["year"]]) for c in cases
                 if str(self.rows[c][COL["year"]] or "").strip().isdigit()]
        case_rows = []
        for c in cases:
            r = self.rows[c]
            mine = [s["role"] for s in seats if s["case"] == c]
            ops = [f"{o['kind']} ({o['decision']})"
                   for o in self.opinions.get(c, []) if o["token"] == token]
            case_rows.append(dict(
                case=r[COL["short"]], year=r[COL["year"]], role="/".join(mine),
                rules=str(r[COL["rules"]] or "").split("(")[0].strip(),
                status=r[COL["status"]],
                breaches_found=str(r[COL["found"]] or ""),
                my_opinion="; ".join(ops),
                italaw=str(r[COL["italaw"]] or "").strip()))
        # v0.9.3 (A6): served vs replaced split for the headline line.
        seats_replaced = sum(1 for s in seats if "(replaced)" in s["role"])
        return dict(
            token=token, seats=len(seats), cases=len(cases),
            seats_replaced=seats_replaced,
            seats_served=len(seats) - seats_replaced,
            has_wing_seat=bool(wing_cases),
            roles=dict(roles), appointing=dict(appointing),
            years=(min(years), max(years)) if years else None,
            pending_at_snapshot=pending,
            opinions=opinions_by_me, opinions_known_denominator=opinions_known,
            possible_unattributed_opinions=stray_opinions,
            annulment_committees=self.annulment_by_token.get(token, 0),
            rules=dict(col_counter(COL["rules"]).most_common(6)),
            institutions=dict(col_counter(COL["inst"]).most_common(4)),
            sectors=dict(col_counter(COL["sector"]).most_common(5)),
            respondents=dict(col_counter(COL["respondent"]).most_common(5)),
            home_states=dict(col_counter(COL["home"]).most_common(5)),
            iias=dict(col_counter(COL["iia"]).most_common(5)),
            iias_total=len(col_counter(COL["iia"])),
            statuses=dict(col_counter(COL["status"]).most_common(8)),
            issue_alleged=dict(issue_alleged.most_common(10)),
            issue_found=dict(issue_found.most_common(10)),
            opinion_in_found_cases=dict(opinion_in_found_cases),
            co_panelists=dict(co.most_common(6)),
            sat_under_presidents=dict(presidents.most_common(4)),
            case_rows=case_rows)

    # ---------- shortlist ----------
    def shortlist(self, issue=None, found_only=False, rules=None, sector=None,
                  respondent=None, role=None, treaty=None, weights=None, top=10,
                  min_seats=3, respondent_gate=False, raw_scores=False):
        weights = weights or {}
        cands = Counter(s["token"] for s in self.seats)
        # v0.9.1 (A3/F10): --treaty — substring filter over the APPLICABLE IIA
        # column. Acts as a gate (zero treaty cases -> screened out) and adds a
        # treaty_match subscore, so e.g. ECT work no longer relies on the
        # sector proxy.
        treaty_cases = None
        if treaty:
            treaty_cases = {i for i, r in enumerate(self.rows)
                            if norm(treaty) in norm(str(r[COL["iia"]] or ""))}
        results = []
        for token, n in cands.items():
            if n < min_seats or token == "Name not available":
                continue
            p = self.profile(token)
            sub = {}
            if treaty_cases is not None:
                my_cases = {s["case"] for s in self.seats if s["token"] == token}
                sub["treaty_match"] = len(my_cases & treaty_cases)
                if sub["treaty_match"] == 0:
                    continue  # no appointment under the requested treaty
            if issue:
                # v0.9.5 (E2): --issue is repeatable; a category counts once
                # even if it matches several of the requested substrings.
                issues = [issue] if isinstance(issue, str) else list(issue)
                key_a = sum(v for k, v in p["issue_alleged"].items()
                            if any(norm(i) in norm(k) for i in issues))
                key_f = sum(v for k, v in p["issue_found"].items()
                            if any(norm(i) in norm(k) for i in issues))
                sub["issue_alleged"] = key_a
                sub["issue_found"] = key_f
                if found_only and key_f == 0:
                    continue
            if rules:
                sub["rules_match"] = sum(v for k, v in p["rules"].items() if norm(rules) in norm(k))
            if sector:
                sub["sector_match"] = sum(v for k, v in p["sectors"].items() if norm(sector) in norm(k))
            if respondent:
                sub["respondent_match"] = sum(v for k, v in p["respondents"].items()
                                              if norm(respondent) in norm(k))
                # v0.9.3 (A18): optional gating — at the default weight (0.5,
                # non-gating) the respondent filter is effectively inert for
                # State-specific searches; --respondent-gate screens out
                # candidates with no appointments in cases against that State.
                if respondent_gate and sub["respondent_match"] == 0:
                    continue
            if role:
                # v0.9.2 (A16): accept the documentation's own vocabulary, not
                # only the raw UNCTAD labels.
                syn = ROLE_SYNONYMS.get(norm(role))
                if syn:
                    sub["role_fit"] = sum(v for k, v in p["roles"].items()
                                          if any(t in norm(k) for t in syn))
                else:
                    sub["role_fit"] = sum(v for k, v in p["roles"].items() if norm(role) in norm(k))
                if sub["role_fit"] == 0:
                    continue  # requested role never held at snapshot — screened out
            sub["experience"] = p["seats"]
            sub["recency"] = sum(1 for cr in p["case_rows"]
                                 if str(cr["year"]).isdigit() and int(cr["year"]) >= 2016)
            sub["availability_penalty"] = p["pending_at_snapshot"]
            results.append(dict(token=token, subscores=sub,
                                seats=p["seats"], pending=p["pending_at_snapshot"],
                                roles=p["roles"]))
        # v0.9.6 (E1): sub-scores are raw counts on wildly different scales,
        # so raw weighting lets volume dominate an ostensibly issue-weighted
        # ranking (121 role-fit appointments at weight 2.0 outscore 8 issue
        # findings at weight 3.0 by 242 to 24). Default scoring therefore
        # normalises each signal to the surviving pool's maximum before its
        # weight is applied — score = SUM(w_k * count_k / poolmax_k) — so a
        # weight means what it says regardless of the signal's scale. Scores
        # are pool-relative (comparable within one query, not across queries);
        # rankings are the output that matters. raw_scores=True restores the
        # legacy raw-count scoring unchanged.
        if raw_scores:
            for r in results:
                score = 0.0
                for k, v in r["subscores"].items():
                    w = weights.get(k, DEFAULT_WEIGHTS.get(k, 0.0))
                    score += w * v
                r["score"] = round(score, 3)
        else:
            poolmax = {}
            for r in results:
                for k, v in r["subscores"].items():
                    poolmax[k] = max(poolmax.get(k, 0), v)
            for r in results:
                contrib, score = {}, 0.0
                for k, v in r["subscores"].items():
                    w = weights.get(k, DEFAULT_WEIGHTS.get(k, 0.0))
                    c = w * (v / poolmax[k]) if poolmax[k] else 0.0
                    contrib[k] = round(c, 2)
                    score += c
                r["contributions"] = contrib
                r["score"] = round(score, 3)
        results.sort(key=lambda x: -x["score"])
        return results[:top]


# v0.9.2 (A16): natural-language role tokens mapped onto UNCTAD's raw labels
# (President / Claimant / Respondent / Sole arbitrator / Unknown, plus
# "(replaced)" variants). "co-arbitrator" — the term the skill's own docs use —
# previously matched nothing and returned an empty shortlist silently.
ROLE_SYNONYMS = {
    "co-arbitrator": ("respondent", "claimant"),
    "coarbitrator": ("respondent", "claimant"),
    "wing": ("respondent", "claimant"),
    "chair": ("president",),
    "sole": ("sole arbitrator",),
}

DEFAULT_WEIGHTS = {  # proposed defaults — printed with every shortlist; user-adjustable
    "issue_found": 3.0,
    "issue_alleged": 1.5,
    "treaty_match": 2.0,
    "rules_match": 1.0,
    "sector_match": 1.0,
    "respondent_match": 0.5,
    "role_fit": 2.0,
    "experience": 0.3,
    "recency": 0.5,
    "availability_penalty": -0.75,
}


ICSID_PROFILE_URL = ("https://icsid.worldbank.org/resources/databases/"
                     "arbitrators-conciliators-ad-hoc-committee-members/profile?cvid={}")
# v0.9.1 (A2/G5/KF2): UNCT/ added — ICSID also administers UNCITRAL cases under
# UNCT/xx/x numbers; the old pattern silently dropped them.
ICSID_CASE_NO_RE = re.compile(r"(?:ARB|CONC|ADM|UNCT)(?:\s*\(AF\))?/\d{2}/\d+")


def load_cvid_map(path):
    if os.path.exists(path):
        with open(path) as f:
            return json.load(f)
    # v0.9.3 (A5/F13): an absent map previously surfaced only as a per-person
    # "no cvid-map match" line, indistinguishable from a genuine no-entry.
    print(f"[warning: cvid map not found at {path} — live-profile lookup runs "
          f"with NO map; every profile will print the no-entry line. Restore "
          f"assets/icsid-cvid-map.json or pass --cvid-map PATH.]", file=sys.stderr)
    return {}


def cvid_lookup(token, cvid_map, overlay=None):
    """Match a 'Surname, I.' token against map keys like 'STERN, Brigitte'.
    Returns list of (map_name, cvid) candidates.

    v0.9.1 (A1/TC1): two guards added.
    1. Initials rule: ALL of the token's initials must be consistent with the
       map key's given names (token initials a prefix of the map given-name
       initials) — 'Cheng, T.-H.' (TH) no longer matches 'CHENG, Teresa' (T).
    2. Distinct guard: a token on the curated-aliases `distinct` list gets a
       cvid ONLY if `distinct_identities_note` names this exact token with a
       full name that matches the map key. Unverified identity = no match —
       the profile then honestly reports no live-lookup entry instead of
       serving the wrong person's page.
    """
    overlay = overlay or {}
    sn, ini = token_parts(token)

    def given_initials(given):
        return "".join(re.findall(r"[A-Z]", given))

    hits = []
    for name, cvid in cvid_map.items():
        if "," in name:
            msn, given = name.split(",", 1)
        else:
            msn, given = name, ""
        if norm(msn) != sn:
            continue
        gi = given_initials(given)
        # token initials must be a prefix of the map given-name initials
        # (token abbreviates the full given names, never the reverse)
        if ini and not gi.startswith(ini):
            continue
        hits.append((name, cvid))

    # distinct guard: for tokens on the `distinct` list, a map key is dropped
    # when it is CONTESTED (the counterpart token in the pair is also
    # initials-compatible with it) and no distinct_identities_note verifies
    # this token's claim to it. Uncontested keys pass; a note, where present,
    # is authoritative in both directions.
    tok_clean = re.sub(r"\s+", " ", token.strip())
    pairs = [p for p in overlay.get("distinct", []) if tok_clean in p]
    if pairs and hits:
        counterparts = {t for p in pairs for t in p if t != tok_clean}
        notes = overlay.get("distinct_identities_note", {})

        def note_verifies(t, map_name):
            note = notes.get(t, "")
            if not note:
                return None  # no note -> unknown
            full_name = note.split("(")[0].strip()
            note_words = {w for w in re.split(r"[^\w]+", norm(full_name)) if len(w) > 1}
            key_words = {w for w in re.split(r"[^\w]+", norm(map_name)) if len(w) > 1}
            return bool(key_words) and key_words.issubset(note_words)

        kept = []
        for name, cvid in hits:
            v = note_verifies(tok_clean, name)
            if v is True:
                kept.append((name, cvid)); continue
            if v is False:
                continue  # note exists and names someone else
            # no note for this token: keep only if no counterpart could
            # equally claim this key on the initials rule
            _, given = name.split(",", 1) if "," in name else (name, "")
            gi = given_initials(given)
            contested = False
            for ct in counterparts:
                _, cini = token_parts(ct)
                if not cini or gi.startswith(cini):
                    contested = note_verifies(ct, name) is not False
                    if contested:
                        break
            if not contested:
                kept.append((name, cvid))
        return kept
    return hits


def icsid_case_numbers(ds, cases):
    """Extract ICSID case numbers for a set of case indices from the Excel."""
    nums = {}
    for c in cases:
        r = ds.rows[c]
        blob = " ".join(str(r[COL[k]] or "") for k in
                        ("full", "decisions", "background", "fo_decisions"))
        m = ICSID_CASE_NO_RE.search(blob)
        if m:
            nums[re.sub(r"\s+", "", m.group(0))] = r[COL["short"]]
    return nums


def delta_report(ds, token, profile_case_numbers, profile_lines=None):
    """Diff the arbitrator's ICSID-profile case list against the Excel.
    profile_case_numbers: iterable of ICSID case-number strings.
    profile_lines: the raw input lines (v0.9.1, A2/S1) — used to (a) count and
    list lines whose case number was NOT recognised, so nothing is dropped
    silently, and (b) offer conservative NAME-match candidates for unnumbered
    profile rows, which previously slipped past the diff entirely."""
    seats = [s for s in ds.seats if s["token"] == token]
    cases = sorted({s["case"] for s in seats})
    excel_nums = icsid_case_numbers(ds, cases)
    prof = {re.sub(r"\s+", "", n) for n in profile_case_numbers}
    in_both = sorted(prof & set(excel_nums))
    profile_only = sorted(prof - set(excel_nums))
    excel_only = sorted(set(excel_nums) - prof)
    pending_at_snap = {n for n, short in excel_nums.items()
                       if str(ds.rows[[c for c in cases
                                       if ds.rows[c][COL["short"]] == short][0]]
                              [COL["status"]] or "").strip() == "Pending"}
    # v0.9.1: account for every input line
    unrecognised = []
    if profile_lines is not None:
        for ln in profile_lines:
            ln = ln.strip()
            if not ln or ln.startswith("#"):
                continue
            if not ICSID_CASE_NO_RE.search(ln):
                unrecognised.append(ln)
    name_candidates = []
    if unrecognised:
        # conservative name matching against THIS arbitrator's Excel cases and
        # the whole corpus: candidate only, never auto-classified
        for ln in unrecognised:
            n_ln = norm(ln)
            for i, r in enumerate(ds.rows):
                short = str(r[COL["short"]] or "")
                if not short:
                    continue
                lead = norm(short.split(" v.")[0])
                resp = norm(str(r[COL["respondent"]] or ""))
                if lead and lead in n_ln and (not resp or resp.split(",")[0] in n_ln):
                    name_candidates.append((ln, short, "this arbitrator's row" if i in cases
                                            else "corpus row (NOT attributed to them in the Excel)"))
                    break
    L = [f"## Newer-cases check — {token}"]
    L.append(f"Excel cases with an ICSID number: {len(excel_nums)} of {len(cases)} total appointments")
    if profile_lines is not None:
        n_lines = sum(1 for ln in profile_lines if ln.strip() and not ln.strip().startswith('#'))
        L.append(f"Input lines supplied: {n_lines}; case numbers recognised: {len(prof)}; "
                 f"lines with no recognisable case number: {len(unrecognised)}")
    L.append(f"ICSID profile case numbers supplied: {len(prof)}")
    L.append(f"- In both: {len(in_both)}")
    L.append(f"- Profile-only (absent from 2023 Excel — post-snapshot, non-treaty, or "
             f"annulment-phase): {len(profile_only)} -> {', '.join(profile_only) or '—'}")
    L.append(f"- Excel-only (non-ICSID fora, or number not on profile): {len(excel_only)}")
    L.append(f"- Pending at snapshot (need live status check): {len(pending_at_snap)} -> "
             f"{', '.join(sorted(pending_at_snap)) or '—'}")
    n_new = len(profile_only) + len(pending_at_snap)
    branch = ("<=6: download each newly decided case and check issues directly"
              if n_new <= 6 else
              ">6: websearch reports to SELECT cases, then targeted downloads — "
              "NO issue-outcome claim enters a table without the downloaded decision")
    L.append(f"Delta workflow branch (threshold 6, candidates to check: {n_new}): {branch}")
    # v0.9.2 (A13): the case-number year is the REGISTRATION year, not the
    # appointment date. A 2023-numbered case can hold a 2024 appointment and
    # vice versa (proven: ARB/23/43 accepted 02/2024 vs ARB/23/25 accepted
    # 11/2023). Flag boundary-vintage profile-only cases for per-case
    # acceptance-date classification.
    snap_year = int(SNAPSHOT.split("/")[-1])
    boundary = [n for n in profile_only
                for m in [re.search(r"/(\d{2})/", n)]
                if m and 2000 + int(m.group(1)) <= snap_year]
    if boundary:
        L.append("BOUNDARY-VINTAGE CASES (case-number year <= snapshot year): "
                 + ", ".join(boundary))
        L.append("  ! Case-number year is the REGISTRATION year, not the appointment "
                 "date. For each case above, fetch the ICSID case page and classify "
                 "by this arbitrator's ACCEPTANCE-OF-APPOINTMENT date against the "
                 "baseline — do not classify from the number alone (v0.9.2/A13).")
    L.append("CLASSIFICATION RULE (v0.9.2): 'new since the cutoff' is governed by the "
             "arbitrator's acceptance date where the case page records one; absence "
             "from the snapshot dataset is reported as its own class, not conflated "
             "with post-cutoff appointment.")
    if unrecognised:
        L.append("UNRECOGNISED INPUT LINES (no case number parsed — check these BY NAME "
                 "against the profile page; do not drop them):")
        for ln in unrecognised:
            L.append(f"  ! {ln[:120]}")
    if name_candidates:
        L.append("Possible name matches for unrecognised lines (CANDIDATES ONLY — verify "
                 "against the profile page and the Excel row before classifying):")
        for ln, short, where in name_candidates:
            L.append(f"  ? {ln[:60]} ~ {short} [{where}]")
    L.append("NOTE: which of the profile-only / pending cases are actually CONCLUDED "
             "requires the live ICSID status per case; this report only stages the check.")
    L.append("NOTE: the ICSID profile page is a FLOOR, not an exhaustive case list "
             "(coverage gaps measured in testing); absence of a case from the profile "
             "is weak evidence of inactivity.")
    return "\n".join(L)


def write_annex(ds, p, path):
    """R4: full per-case table as xlsx with a Sources sheet."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cases"
    # v0.9.1 (A8/R5): column spec per Green's ruling of 2026-08-04.
    hdr = ["Case name", "Case number", "Year initiated", "Appointed by", "Role",
           "Rules", "Case status (tribunal-level)", "Substantive decisions",
           "Separate opinions", "italaw case page"]
    ws.append(hdr)

    def split_role(role_label):
        """UNCTAD's single label conflates appointing side and role.
        'X (replaced)' -> the '(replaced)' qualifier stays on the Role."""
        repl = " (replaced)" if "(replaced)" in role_label else ""
        base = role_label.replace(" (replaced)", "").strip()
        if base in ("Respondent", "Claimant"):
            return base, "Co-arbitrator" + repl
        if base == "President":
            return "Not recorded in dataset", "President" + repl
        if base == "Sole arbitrator":
            return "Not recorded in dataset", "Sole arbitrator" + repl
        return "Not recorded in dataset", (base + repl if base else "Not recorded")

    cases_by_short = {}
    for s in ds.seats:
        if s["token"] == p["token"]:
            cases_by_short[ds.rows[s["case"]][COL["short"]]] = s["case"]
    for cr in p["case_rows"]:
        c = cases_by_short.get(cr["case"])
        subs = "; ".join(substantive_decisions(ds.rows[c][COL["decisions"]])) if c is not None else ""
        icsid_no = ""
        if c is not None:
            m = ICSID_CASE_NO_RE.search(" ".join(
                str(ds.rows[c][COL[k]] or "") for k in ("full", "decisions", "background")))
            icsid_no = m.group(0) if m else ""
        appointed_by, role_out = split_role(cr["role"])
        ws.append([cr["case"], icsid_no, cr["year"], appointed_by, role_out,
                   cr["rules"], cr["status"], subs, cr["my_opinion"], cr["italaw"]])
    src = wb.create_sheet("Sources")
    # v0.9.2 (A12): the template requires the work-product header on EVERY
    # deliverable including the annex Sources sheet — the re-run graded its
    # absence on four items.
    src.append(["Attorney work product — prepared at counsel's direction for "
                "arbitrator selection"])
    src.append(["All rows: UNCTAD ISDS Navigator full-data Excel, snapshot "
                f"{SNAPSHOT} (user's own download; treaty-based ISDS only; "
                "not exhaustive; tribunal-level outcomes, not personal votes)."])
    src.append(["'Year initiated' is the year the CASE was initiated (filed) per UNCTAD — "
                "NOT the year of this arbitrator's appointment, which the dataset does not "
                "record and which can be later (appointments to pre-existing cases)."])
    src.append(["'Appointed by' is derived from UNCTAD's role label: wing members carry the "
                "appointing side; for presidents and sole arbitrators the appointing party "
                "(parties / Chairman / Secretary-General) is not recorded in the dataset — "
                "take it from the ICSID profile or case page where retrieved."])
    src.append(["'Separate opinions' lists published individual opinions by THIS arbitrator "
                "recorded for the case; silence is not proof of unanimity."])
    src.append([f"Live Navigator last known: {LAST_KNOWN_NAVIGATOR['total']} cases "
                f"as of {LAST_KNOWN_NAVIGATOR['as_of']}."])
    src.append(["Not legal advice; research aid only."])

    # v0.9.4: presentation formatting — spec mirrored in
    # references/output-template.md "Annex formatting (xlsx)"; ad-hoc sheets
    # built without this engine must match it.
    n_rows, n_cols = ws.max_row, len(hdr)
    hdr_fill = PatternFill("solid", fgColor="DDE3EA")
    band_fill = PatternFill("solid", fgColor="F5F7FA")
    wrap_top = Alignment(wrap_text=True, vertical="top")
    plain_top = Alignment(vertical="top")
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.fill = hdr_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for i, w in enumerate([45, 15, 16, 22, 20, 12, 26, 48, 40, 42], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    wrap_cols = {1, 4, 5, 7, 8, 9}
    for r in range(2, n_rows + 1):
        banded = (r % 2 == 0)
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            if c == 10 and isinstance(cell.value, str) and cell.value.startswith("http"):
                cell.hyperlink = cell.value  # clickable italaw link
                cell.style = "Hyperlink"     # sets blue/underline; reapply layout after
            cell.alignment = wrap_top if c in wrap_cols else plain_top
            if banded:
                cell.fill = band_fill
        ws.cell(row=r, column=1).font = Font(italic=True)  # case names italicised
    ws.freeze_panes = "A2"
    if n_rows > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{n_rows}"
    src.column_dimensions["A"].width = 110
    for r in range(1, src.max_row + 1):
        src.cell(row=r, column=1).alignment = wrap_top
    src.cell(row=1, column=1).font = Font(bold=True)
    wb.save(path)


def freshness_footer():
    return (
        "\n---\nDATA FRESHNESS: computed from the UNCTAD ISDS Navigator full-data Excel, "
        f"snapshot {SNAPSHOT} ({SNAPSHOT_CASES} cases). The live Navigator reported "
        f"{LAST_KNOWN_NAVIGATOR['total']} cases as of {LAST_KNOWN_NAVIGATOR['as_of']} "
        f"(last observed {LAST_KNOWN_NAVIGATOR['observed']}). Counts above EXCLUDE anything "
        "after the snapshot — for active arbitrators they materially understate current "
        "activity; run the live ICSID profile lookup and the newer-cases check before "
        "relying on them (and note the profile page is a floor, not a case list).\n"
        "SCOPE: treaty-based ISDS only; commercial-arbitration appointments and confidential "
        "cases are invisible; UNCTAD data cannot be deemed exhaustive. Fuller analytics exist "
        "in subscription databases (Jus Mundi, GAR ART) not used by this tool.\n"
        "Not legal advice; research aid only.\n")


def profile_md(p):
    # v0.9.1 (A7/B1): no builder jargon, "Appointments" never "Seats".
    L = [f"# Dataset profile — {p['token']} (UNCTAD 2023 Excel, snapshot {SNAPSHOT})", ""]
    yrs = f"{p['years'][0]}–{p['years'][1]}" if p["years"] else "n/a"
    # v0.9.3 (A6): served vs replaced split in the headline, printed only when
    # a replaced appointment exists.
    repl = p.get("seats_replaced", 0)
    served_note = f" (served {p['seats_served']}; replaced in {repl})" if repl else ""
    L.append(f"Appointments: **{p['seats']}**{served_note} in {p['cases']} cases | first–last case "
             f"initiation year: {yrs} | pending at snapshot: {p['pending_at_snapshot']} "
             f"| annulment-committee entries: {p['annulment_committees']}")
    if p.get("icsid_profile"):
        for name, cvid in p["icsid_profile"]:
            L.append(f"ICSID profile: **{name}** — {ICSID_PROFILE_URL.format(cvid)} "
                     f"(CONFIRM the page is this person before relying on it; self-reported "
                     f"data; the profile is a FLOOR, not an exhaustive case list)")
    elif p.get("icsid_profile") == []:
        L.append("ICSID profile: no verified entry in the cvid map for this person — do NOT "
                 "guess a profile URL. Offer the ICSID arbitrator listing page via the "
                 "user's browser, or refresh the map from a newer page save. The dataset "
                 "figures below stand on their own.")
    L.append("")
    L.append("| Dimension | Breakdown |")
    L.append("|---|---|")
    L.append(f"| Roles | {p['roles']} |")
    L.append(f"| Appointing side (base role) | {p['appointing']} |")
    L.append(f"| Rules | {p['rules']} |")
    L.append(f"| Sectors (top) | {p['sectors']} |")
    L.append(f"| Respondent States (top) | {p['respondents']} |")
    L.append(f"| Investor home States (top) | {p['home_states']} |")
    # v0.9.3 (A15): the treaty row previously said "(top)" while silently
    # dropping everything past five — now the total is explicit and the full
    # list is pointed to.
    n_iias = p.get("iias_total", len(p["iias"]))
    L.append(f"| Treaties (top {min(5, n_iias)} of {n_iias}; full list via the case annex) | {p['iias']} |")
    L.append(f"| Case outcomes (tribunal-level, NOT personal votes) | {p['statuses']} |")
    L.append(f"| Co-panelists (top) | {p['co_panelists']} |")
    # v0.9.3 (A6): row restricted to wing appointments; suppressed when the
    # subject held no wing seat (the old row listed successor presidents /
    # peers in those cases).
    if p.get("sat_under_presidents"):
        L.append(f"| Presidents sat under (wing appointments only) | {p['sat_under_presidents']} |")
    L.append("")
    n_op = len(p["opinions"])
    L.append(f"## Published individual opinions: {n_op} "
             f"(denominator: {p['opinions_known_denominator']} cases with opinions data known)")
    L.append("Stated assumption: silence in a known row = no published individual opinion; "
             "unanimity is not thereby proven. Per-holding votes require retrieving the "
             "decisions themselves (\"check the record\").")
    for o in p["opinions"]:
        lang = f" [languages: {', '.join(o['languages'])}]" if o.get("languages") else ""
        L.append(f"- {o['case']}: {o['kind']} — {o['decision']}{lang} "
                 f"(recorded author: {o['author']})")
    # v0.9.3 (A10/G4): dataset opinions naming this arbitrator in cases whose
    # ARBITRATORS field does not attribute them (e.g. "Data not available")
    # are invisible to the count above — surface, never silently add.
    if p.get("possible_unattributed_opinions"):
        L.append("")
        L.append("! Possible additional opinion(s) recorded in case(s) NOT attributed to "
                 "this arbitrator in the dataset's arbitrator field — excluded from the "
                 "count above; verify identity on the case page before any reliance:")
        for o in p["possible_unattributed_opinions"]:
            L.append(f"  - {o['case']}: {o['kind']} — {o['decision']} "
                     f"(recorded author: {o['author']})")
    L.append("")
    L.append("## Issue exposure (tribunal-level; NEVER a personal win-rate)")
    # v0.9.3 (A4/F12): column renamed and disclaimed — it counts co-occurrence
    # (an opinion of any kind in a case where that breach was found), not a
    # dissent on that issue.
    L.append("| Issue (breach FOUND) | Sat on tribunal, breach found | Of which, they published an individual opinion§ |")
    L.append("|---|---|---|")
    for k, v in p["issue_found"].items():
        L.append(f"| {k} | {v} | {p['opinion_in_found_cases'].get(k, 0)} |")
    L.append("§ co-occurrence only: an individual opinion of ANY kind in a case where "
             "that breach was found — NOT a dissent on that issue. Per-issue votes "
             "require the decisions themselves (\"check the record on her/his votes\").")
    L.append("")
    L.append("[Record checks not run: per-holding votes, law firms, challenges, "
             "double-hatting, publications review — available on request "
             "(\"check the record on <topic>\")]")
    return "\n".join(L)


def find_excel(preferred):
    """v0.9.2 (A14): locate the user's UNCTAD full-data workbook.

    The path is no longer hard-coded to one exact filename: if the preferred
    path is absent, any .xlsx in the skill's data/ folder (then the current
    directory) is considered, and validated by opening it and checking the
    header row for the UNCTAD schema. Exactly one validated candidate is used
    (with a notice); zero or several produce actionable guidance instead of a
    bare FileNotFoundError."""
    import glob
    if os.path.exists(preferred):
        return preferred

    def is_unctad(path):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True)
            ws = wb[wb.sheetnames[0]]
            for r in ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW,
                                  values_only=True):
                cells = [str(c or "") for c in r]
                return "ARBITRATORS" in cells and "NO." in cells
        except Exception:
            return False
        return False

    search_dirs = [os.path.dirname(preferred) or ".", "."]
    candidates = []
    for d in search_dirs:
        for f in sorted(glob.glob(os.path.join(d, "*.xlsx"))):
            if f not in candidates and is_unctad(f):
                candidates.append(f)
    if len(candidates) == 1:
        print(f"[excel not found at {preferred}; using validated UNCTAD workbook "
              f"{candidates[0]} instead (v0.9.2 discovery)]", file=sys.stderr)
        return candidates[0]
    if len(candidates) > 1:
        sys.exit(f"ERROR: no workbook at {preferred} and several UNCTAD-schema "
                 f"candidates found: {candidates}. Pass one explicitly with --excel.")
    sys.exit(
        f"ERROR: the UNCTAD ISDS Navigator full-data Excel was not found.\n"
        f"  Looked for: {preferred}, then any *.xlsx with the UNCTAD schema in "
        f"{search_dirs}.\n"
        f"  Fix: download it yourself (investmentpolicy.unctad.org -> ISDS "
        f"Navigator -> full data download) and place it in the skill's data/ "
        f"folder (any filename), or pass --excel PATH.\n"
        f"  NOTE (cloud sessions): the container-side data/ folder does not "
        f"persist between sessions — check the connected project folder (e.g. "
        f"isds-arbitrators/data/) for the user's copy before asking for a new "
        f"upload.")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--excel", default=os.path.join(os.path.dirname(__file__) or ".",
                                                    "..", "data",
                                                    "UNCTAD-ISDS-Navigator-data-set-31December2023.xlsx"))
    ap.add_argument("--aliases", default=os.path.join(os.path.dirname(__file__) or ".",
                                                      "..", "assets", "curated-aliases.json"))
    ap.add_argument("--arbitrator")
    ap.add_argument("--list-matches")
    ap.add_argument("--shortlist", action="store_true")
    ap.add_argument("--issue", action="append",
                    help="issue substring; repeatable — repeated flags are combined "
                         "(a case category counts once even if it matches several)")
    ap.add_argument("--rules"); ap.add_argument("--sector")
    ap.add_argument("--respondent"); ap.add_argument("--role")
    ap.add_argument("--raw-scores", action="store_true",
                    help="legacy scoring: weights applied to raw counts with no "
                         "pool normalisation (pre-v0.9.6 behaviour)")
    ap.add_argument("--respondent-gate", action="store_true",
                    help="v0.9.3 (A18): screen OUT candidates with zero appointments "
                         "in cases against --respondent (default: non-gating subscore)")
    ap.add_argument("--treaty", help="substring filter over the APPLICABLE IIA column "
                                     "(v0.9.1, e.g. \"Energy Charter\"); gates the "
                                     "shortlist and adds a treaty_match subscore")
    ap.add_argument("--found", action="store_true", help="require the issue to have been FOUND")
    ap.add_argument("--weights", help="JSON file overriding default weights")
    ap.add_argument("--top", type=int, default=10)
    ap.add_argument("--min-seats", type=int, default=3)
    ap.add_argument("--json", action="store_true")
    ap.add_argument("--audit-unattributed", action="store_true",
                    help="list opinion authors that could not be matched to a tribunal token")
    ap.add_argument("--cvid-map", default=os.path.join(os.path.dirname(__file__) or ".",
                                                       "..", "assets", "icsid-cvid-map.json"))
    ap.add_argument("--annex", help="write the full per-case table to this .xlsx path (R4)")
    ap.add_argument("--delta-file",
                    help="R5: text file containing the arbitrator's ICSID-profile case "
                         "numbers (any format; ARB/xx/x patterns are extracted)")
    args = ap.parse_args()

    overlay = {}
    if os.path.exists(args.aliases):
        with open(args.aliases) as f:
            overlay = json.load(f)
    excel_path = find_excel(args.excel)
    ds = Dataset(excel_path, overlay)
    # v0.9.2 (A14): standard load-verification line (to stderr, so markdown
    # output stays clean). Two 2026-08-04 run logs recorded whole-sheet
    # dimensions instead of the data table; this stops the drift.
    print(f"[dataset loaded: {excel_path} — {len(ds.rows)} cases x "
          f"{len(COL)} data columns; snapshot {SNAPSHOT}]", file=sys.stderr)

    if args.audit_unattributed:
        for case, author in ds.unattributed:
            print(f"UNATTRIBUTED: {case}: {author}")
        print(f"total unattributed: {len(ds.unattributed)}")
        return

    if args.list_matches:
        for t in ds.resolve(args.list_matches):
            print(t)
        return

    if args.arbitrator:
        matches = ds.resolve(args.arbitrator)
        if not matches:
            print(f"No match for {args.arbitrator!r}."); sys.exit(1)
        if len(matches) > 1:
            print(f"Ambiguous — {len(matches)} matches (pick one, exact):")
            for t in matches:
                print("  ", t)
            sys.exit(2)
        p = ds.profile(matches[0])
        p["icsid_profile"] = cvid_lookup(matches[0], load_cvid_map(args.cvid_map), overlay)
        if args.delta_file:
            with open(args.delta_file) as f:
                lines = f.read().splitlines()
            nums = [n for ln in lines for n in ICSID_CASE_NO_RE.findall(ln)]
            print(delta_report(ds, matches[0], nums, profile_lines=lines))
            print()
        if args.annex:
            write_annex(ds, p, args.annex)
            print(f"[annex written: {args.annex}]")
        if args.json:
            print(json.dumps(p, indent=2, ensure_ascii=False, default=str))
        else:
            print(profile_md(p))
            print(freshness_footer())
        return

    if args.shortlist:
        weights = dict(DEFAULT_WEIGHTS)
        if args.weights:
            with open(args.weights) as f:
                weights.update(json.load(f))
        res = ds.shortlist(issue=args.issue, found_only=args.found, rules=args.rules,
                           sector=args.sector, respondent=args.respondent, role=args.role,
                           treaty=args.treaty, weights=weights, top=args.top,
                           min_seats=args.min_seats, respondent_gate=args.respondent_gate,
                           raw_scores=args.raw_scores)
        if args.role and not res:
            print("WARNING: --role %r matched no candidates. Accepted values: the "
                  "UNCTAD labels (President, Claimant, Respondent, Sole arbitrator, "
                  "Unknown) or synonyms co-arbitrator/wing, chair, sole (v0.9.2/A16)."
                  % args.role, file=sys.stderr)
        if args.json:
            # v0.9.2 (A17): --shortlist previously ignored --json entirely.
            print(json.dumps({"filters": dict(issue=args.issue, found_only=args.found,
                                              treaty=args.treaty, rules=args.rules,
                                              sector=args.sector,
                                              respondent=args.respondent,
                                              respondent_gate=args.respondent_gate,
                                              role=args.role,
                                              min_seats=args.min_seats),
                              "weights": weights, "raw_scores": args.raw_scores,
                              "results": res},
                             indent=2, ensure_ascii=False, default=str))
            return
        print("## Shortlist (dataset screening only — per-candidate live check, sourced "
              "availability check, and record checks required before use)")
        print(f"Filters: issue={args.issue!r} found_only={args.found} treaty={args.treaty!r} "
              f"rules={args.rules!r} sector={args.sector!r} respondent={args.respondent!r} "
              f"role={args.role!r} min_seats={args.min_seats}")
        print(f"RANKING WEIGHTS (user-adjustable via --weights): {json.dumps(weights)}")
        if args.raw_scores:
            print("SCORING: legacy raw-count mode (--raw-scores) — weights applied to "
                  "unnormalised counts; volume can dominate the stated weights.")
        else:
            print("SCORING: each signal is normalised to the pool maximum before its "
                  "weight is applied (score = SUM(weight x count/poolmax)), so weights "
                  "mean what they say across signals of different scales. Raw counts "
                  "and each weighted contribution are both shown. Legacy raw-count "
                  "scoring: --raw-scores.")
        print()
        # v0.9.2: "Appointments", never "Seats" (R4 wording rule; raw output
        # leaks into run folders).
        if args.raw_scores:
            print("| # | Arbitrator | Score | Appointments | Pending@snap | Subscores |")
            print("|---|---|---|---|---|---|")
            for i, r in enumerate(res, 1):
                print(f"| {i} | {r['token']} | {r['score']} | {r['seats']} | {r['pending']} "
                      f"| {json.dumps(r['subscores'])} |")
        else:
            print("| # | Arbitrator | Score | Appointments | Pending@snap | Subscores (raw) "
                  "| Weighted contributions |")
            print("|---|---|---|---|---|---|---|")
            for i, r in enumerate(res, 1):
                print(f"| {i} | {r['token']} | {r['score']} | {r['seats']} | {r['pending']} "
                      f"| {json.dumps(r['subscores'])} | {json.dumps(r['contributions'])} |")
        print(freshness_footer())
        return

    ap.print_help()


if __name__ == "__main__":
    main()
